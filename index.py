# -*- coding: utf-8 -*-
"""
Validador de Imagens de Produtos
--------------------------------
- Le o Excel 'Produtos_comercializavel_S.xlsx' (colunas PRO_CODIGO e PRO_DESCRICAO)
- Casa cada produto com a imagem em 'imagens/<PRO_CODIGO>.jpg'
- Abre uma tela web onde voce valida se a imagem BATE ou NAO BATE com a descricao
- Grava o resultado no PostgreSQL (connection string na variavel DATABASE do .env)

Como usar:
    .venv\\Scripts\\python.exe index.py
e abra http://127.0.0.1:5000 no navegador.
"""

import hashlib
import os
from functools import wraps

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from flask import Flask, abort, g, jsonify, redirect, request, send_from_directory, session, url_for

# ----------------------------------------------------------------------------
# Configuracao
# ----------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE_DIR, "Produtos_comercializavel_S.xlsx")
PASTA_IMAGENS = os.path.join(BASE_DIR, "imagens")
SQLITE_ANTIGO = os.path.join(BASE_DIR, "banco.db")  # so para importar dados antigos

load_dotenv(os.path.join(BASE_DIR, ".env"))
DATABASE = os.environ.get("DATABASE", "").strip()

# Credenciais: USUARIOS=carlos:senha123,renato:senha456
_usuarios_raw = os.environ.get("USUARIOS", "carlos:carlos,renato:renato")
USUARIOS = {}
for _par in _usuarios_raw.split(","):
    _par = _par.strip()
    if ":" in _par:
        _u, _s = _par.split(":", 1)
        USUARIOS[_u.strip().lower()] = _s.strip()

# Status possiveis
PENDENTE = "pendente"
APROVADO = "aprovado"   # imagem BATE com a descricao
REPROVADO = "reprovado"  # imagem NAO BATE

app = Flask(__name__)
# Chave derivada do DATABASE para sessoes persistirem entre reinicializacoes
app.secret_key = os.environ.get("SECRET_KEY") or hashlib.sha256(DATABASE.encode()).hexdigest()


# ----------------------------------------------------------------------------
# Auth
# ----------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def _redirect_usuario(usuario):
    if usuario == "carlos":
        return redirect(url_for("rota_carlos"))
    if usuario == "renato":
        return redirect(url_for("rota_renato"))
    return redirect(url_for("index"))


LOGIN_HTML = """<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Login — Validador de Imagens</title>
<style>
  :root { --bg:#0f172a; --card:#1e293b; --txt:#e2e8f0; --mut:#94a3b8; --ok:#16a34a; --err:#dc2626; }
  * { box-sizing: border-box; }
  body { margin:0; font-family: system-ui, Segoe UI, sans-serif; background:var(--bg); color:var(--txt); min-height:100vh; display:flex; align-items:center; justify-content:center; }
  .box { background:var(--card); border-radius:16px; padding:36px 32px; width:100%; max-width:360px; box-shadow:0 20px 60px rgba(0,0,0,.4); }
  h1 { margin:0 0 6px; font-size:22px; }
  p { margin:0 0 24px; color:var(--mut); font-size:14px; }
  label { display:block; font-size:13px; color:var(--mut); margin-bottom:6px; }
  input { width:100%; padding:10px 14px; border-radius:10px; border:1px solid #334155; background:#0f172a; color:var(--txt); font-size:15px; margin-bottom:16px; outline:none; transition:border-color .15s; }
  input:focus { border-color:#38bdf8; }
  button[type=submit] { width:100%; padding:13px; border:0; border-radius:10px; background:var(--ok); color:#fff; font-size:16px; font-weight:700; cursor:pointer; }
  button[type=submit]:hover { opacity:.9; }
  .erro { background:#450a0a; border:1px solid var(--err); color:#fca5a5; border-radius:8px; padding:10px 14px; font-size:14px; margin-bottom:16px; }
</style>
</head>
<body>
<div class="box">
  <h1>&#128269; Validador de Imagens</h1>
  <p>Entre com suas credenciais para continuar.</p>
  {erro_html}
  <form method="post" action="/login">
    <label>Usu&#225;rio</label>
    <input name="usuario" type="text" autocomplete="username" autofocus placeholder="seu usu&#225;rio">
    <label>Senha</label>
    <input name="senha" type="password" autocomplete="current-password" placeholder="&#8226;&#8226;&#8226;&#8226;&#8226;&#8226;&#8226;&#8226;">
    <button type="submit">Entrar</button>
  </form>
</div>
</body>
</html>"""


# ----------------------------------------------------------------------------
# Banco de dados
# ----------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = psycopg2.connect(DATABASE)
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def db_query(sql, params=()):
    """SELECT: retorna lista de dicts."""
    db = get_db()
    with db.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params)
        return cur.fetchall()


def db_execute(sql, params=()):
    """INSERT/UPDATE: executa e commita."""
    db = get_db()
    with db.cursor() as cur:
        cur.execute(sql, params)
    db.commit()


# Ordena numericamente quando o codigo e numero, sem quebrar em codigos com letras
ORDEM_NUMERICA = "(CASE WHEN pro_codigo ~ '^[0-9]+$' THEN pro_codigo::bigint END)"


def init_db():
    con = psycopg2.connect(DATABASE)
    cur = con.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS produtos (
            pro_codigo     TEXT PRIMARY KEY,
            pro_descricao  TEXT,
            imagem         TEXT,            -- nome do arquivo da imagem (ou NULL)
            tem_imagem     INTEGER NOT NULL DEFAULT 0,
            status         TEXT NOT NULL DEFAULT 'pendente',
            observacao     TEXT,
            link           TEXT,            -- preenchido quando o usuario marca NAO BATE
            data_validacao TEXT
        )
        """
    )
    cur.execute("ALTER TABLE produtos ADD COLUMN IF NOT EXISTS link TEXT")
    con.commit()

    # Importacao unica: se o Postgres esta vazio e existe o banco.db antigo,
    # traz os dados (status, observacao, link, data) validados no SQLite
    cur.execute("SELECT COUNT(*) FROM produtos")
    if cur.fetchone()[0] == 0 and os.path.exists(SQLITE_ANTIGO):
        import sqlite3

        scon = sqlite3.connect(SQLITE_ANTIGO)
        scon.row_factory = sqlite3.Row
        rows = scon.execute(
            "SELECT pro_codigo, pro_descricao, imagem, tem_imagem, status, observacao, link, data_validacao FROM produtos"
        ).fetchall()
        scon.close()
        for r in rows:
            cur.execute(
                """
                INSERT INTO produtos (pro_codigo, pro_descricao, imagem, tem_imagem, status, observacao, link, data_validacao)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (pro_codigo) DO NOTHING
                """,
                tuple(r),
            )
        con.commit()
        print(f"Importados {len(rows)} produtos do banco.db antigo para o PostgreSQL.")
    con.close()


def carregar_excel():
    """Le o Excel e insere/atualiza os produtos no banco (sem apagar validacoes ja feitas)."""
    import openpyxl

    # mapa de imagens existentes: codigo -> nome do arquivo
    imagens = {}
    if os.path.isdir(PASTA_IMAGENS):
        for f in os.listdir(PASTA_IMAGENS):
            nome, ext = os.path.splitext(f)
            if ext.lower() in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"):
                imagens[nome.strip()] = f

    wb = openpyxl.load_workbook(EXCEL, read_only=True)
    ws = wb.active

    # Localiza as colunas pelo cabecalho
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx_cod = header.index("PRO_CODIGO")
        idx_desc = header.index("PRO_DESCRICAO")
    except ValueError:
        raise RuntimeError(f"Cabecalho nao encontrado. Colunas lidas: {header}")

    con = psycopg2.connect(DATABASE)
    cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM produtos")
    antes = cur.fetchone()[0]
    vistos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = row[idx_cod]
        desc = row[idx_desc]
        if cod is None:
            continue
        cod = str(cod).strip()
        if not cod or cod in vistos:
            continue  # ignora codigos duplicados no Excel (mantem o primeiro)
        vistos.add(cod)
        desc = "" if desc is None else str(desc).strip()
        arquivo = imagens.get(cod)
        tem = 1 if arquivo else 0
        # UPSERT: insere se novo; se existir, atualiza so os campos do Excel e preserva o status validado
        cur.execute(
            """
            INSERT INTO produtos (pro_codigo, pro_descricao, imagem, tem_imagem, status)
            VALUES (%s,%s,%s,%s,%s)
            ON CONFLICT (pro_codigo) DO UPDATE SET
                pro_descricao = EXCLUDED.pro_descricao,
                imagem        = EXCLUDED.imagem,
                tem_imagem    = EXCLUDED.tem_imagem
            """,
            (cod, desc, arquivo, tem, PENDENTE),
        )
    con.commit()
    cur.execute("SELECT COUNT(*) FROM produtos")
    depois = cur.fetchone()[0]
    con.close()
    wb.close()
    return depois - antes


# ----------------------------------------------------------------------------
# Rotas de autenticacao
# ----------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if "usuario" in session:
        return _redirect_usuario(session["usuario"])
    erro = ""
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip().lower()
        senha = request.form.get("senha", "").strip()
        if usuario in USUARIOS and USUARIOS[usuario] == senha:
            session["usuario"] = usuario
            return _redirect_usuario(usuario)
        erro = "Usu\u00e1rio ou senha inv\u00e1lidos."
    erro_html = f'<div class="erro">{erro}</div>' if erro else ""
    return LOGIN_HTML.replace("{erro_html}", erro_html)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ----------------------------------------------------------------------------
# Rotas da pagina
# ----------------------------------------------------------------------------
def render_pagina(ordem="asc", nome=""):
    import json as _json
    usuario = session.get("usuario", "")
    inj = (f"<script>window.ORDEM={_json.dumps(ordem)};"
           f"window.VALIDADOR={_json.dumps(nome)};"
           f"window.USUARIO_LOGADO={_json.dumps(usuario)};</script>")
    return PAGINA_HTML.replace("<!--INJECT-->", inj)


@app.route("/")
@login_required
def index():
    return render_pagina()


@app.route("/carlos")
@login_required
def rota_carlos():
    # comeca das imagens em ordem CRESCENTE
    return render_pagina(ordem="asc", nome="Carlos")


@app.route("/renato")
@login_required
def rota_renato():
    # comeca das imagens em ordem DECRESCENTE
    return render_pagina(ordem="desc", nome="Renato")


@app.route("/imagens/<path:nome>")
@login_required
def servir_imagem(nome):
    return send_from_directory(PASTA_IMAGENS, nome)


@app.route("/api/stats")
@login_required
def api_stats():
    base = "FROM produtos WHERE tem_imagem = 1"
    total = db_query(f"SELECT COUNT(*) c {base}")[0]["c"]
    aprov = db_query(f"SELECT COUNT(*) c {base} AND status=%s", (APROVADO,))[0]["c"]
    repro = db_query(f"SELECT COUNT(*) c {base} AND status=%s", (REPROVADO,))[0]["c"]
    pend = db_query(f"SELECT COUNT(*) c {base} AND status=%s", (PENDENTE,))[0]["c"]
    return jsonify(total=total, aprovado=aprov, reprovado=repro, pendente=pend)


@app.route("/api/lista")
@login_required
def api_lista():
    """Lista produtos (com imagem) para navegacao, com filtro opcional por status."""
    status = request.args.get("status", "todos")
    sql = "SELECT pro_codigo, pro_descricao, imagem, status, observacao, link FROM produtos WHERE tem_imagem = 1"
    params = []
    if status in (PENDENTE, APROVADO, REPROVADO):
        sql += " AND status = %s"
        params.append(status)
    ordem = "DESC" if request.args.get("ordem") == "desc" else "ASC"
    sql += f" ORDER BY {ORDEM_NUMERICA} {ordem}, pro_codigo {ordem}"
    rows = db_query(sql, params)
    return jsonify([dict(r) for r in rows])


@app.route("/api/status")
@login_required
def api_status():
    """Retorna todos os produtos com pro_codigo, status e link (JSON)."""
    rows = db_query(
        f"SELECT pro_codigo, status, link FROM produtos ORDER BY {ORDEM_NUMERICA}, pro_codigo"
    )
    return jsonify([dict(r) for r in rows])


@app.route("/api/validar", methods=["POST"])
@login_required
def api_validar():
    data = request.get_json(force=True)
    cod = str(data.get("pro_codigo", "")).strip()
    status = data.get("status")
    obs = data.get("observacao")
    link = data.get("link")
    if status not in (APROVADO, REPROVADO, PENDENTE):
        abort(400, "status invalido")
    if status != REPROVADO:
        link = None  # link so faz sentido quando NAO BATE
    elif link is not None:
        link = str(link).strip() or None
    from datetime import datetime
    db_execute(
        "UPDATE produtos SET status = %s, observacao = %s, link = %s, data_validacao = %s WHERE pro_codigo = %s",
        (status, obs, link, datetime.now().isoformat(timespec="seconds"), cod),
    )
    return jsonify(ok=True)


# ----------------------------------------------------------------------------
# Pagina (HTML + CSS + JS em uma string)
# ----------------------------------------------------------------------------
PAGINA_HTML = r"""<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Validador de Imagens</title>
<style>
  :root { --ok:#16a34a; --no:#dc2626; --bg:#0f172a; --card:#1e293b; --txt:#e2e8f0; --mut:#94a3b8; }
  * { box-sizing: border-box; }
  body { margin:0; font-family: system-ui, Segoe UI, sans-serif; background:var(--bg); color:var(--txt); }
  header { padding:12px 20px; background:#020617; display:flex; gap:18px; align-items:center; flex-wrap:wrap; position:sticky; top:0; z-index:5; }
  header h1 { font-size:18px; margin:0; }
  .stats { display:flex; gap:14px; font-size:14px; }
  .stats b { font-weight:700; }
  .pill { padding:2px 10px; border-radius:999px; background:#334155; }
  .pill.ok { background:var(--ok); } .pill.no { background:var(--no); } .pill.pe { background:#475569; }
  select, button { font-size:14px; }
  select { padding:6px 10px; border-radius:8px; border:1px solid #475569; background:var(--card); color:var(--txt); }
  main { max-width:760px; margin:24px auto; padding:0 16px; }
  .card { background:var(--card); border-radius:16px; padding:20px; box-shadow:0 10px 30px rgba(0,0,0,.3); }
  .codigo { font-size:14px; color:var(--mut); }
  .descricao { font-size:22px; font-weight:700; margin:6px 0 16px; line-height:1.3; }
  .imgwrap { background:#fff; border-radius:12px; display:flex; align-items:center; justify-content:center; min-height:340px; overflow:hidden; }
  .imgwrap img { max-width:100%; max-height:460px; object-fit:contain; }
  .badge { display:inline-block; margin-top:12px; padding:4px 12px; border-radius:999px; font-size:13px; font-weight:700; }
  .badge.ok{background:var(--ok)} .badge.no{background:var(--no)} .badge.pe{background:#475569}
  .acoes { display:flex; gap:12px; margin-top:18px; }
  .acoes button { flex:1; padding:16px; border:0; border-radius:12px; font-size:17px; font-weight:700; cursor:pointer; color:#fff; }
  .btn-ok { background:var(--ok); } .btn-no { background:var(--no); }
  .nav { display:flex; justify-content:space-between; align-items:center; margin-top:14px; gap:10px; }
  .nav button { padding:10px 16px; border-radius:10px; border:1px solid #475569; background:var(--card); color:var(--txt); cursor:pointer; }
  .nav button:disabled { opacity:.4; cursor:default; }
  .pos { color:var(--mut); font-size:14px; }
  .dica { text-align:center; color:var(--mut); font-size:12px; margin-top:14px; }
  .vazio { text-align:center; padding:60px 20px; color:var(--mut); }
  .btn-sair { padding:6px 14px; border-radius:8px; border:1px solid #334155; background:transparent; color:var(--mut); cursor:pointer; font-size:13px; white-space:nowrap; }
  .btn-sair:hover { border-color:#475569; color:var(--txt); }
</style>
</head>
<body>
<!--INJECT-->
<header>
  <h1>🔍 Validador de Imagens <span id="quem" style="color:#38bdf8"></span></h1>
  <div class="stats" id="stats"></div>
  <div style="margin-left:auto; display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
    <label class="pos">Mostrar:</label>
    <select id="filtro">
      <option value="pendente">Pendentes</option>
      <option value="todos">Todos</option>
      <option value="aprovado">Aprovados (bate)</option>
      <option value="reprovado">Reprovados (nao bate)</option>
    </select>
    <span id="usuario-logado" style="color:#64748b; font-size:13px;"></span>
    <button class="btn-sair" id="btn-sair">Sair</button>
  </div>
</header>

<main id="main">
  <div class="card" id="card" style="display:none">
    <div class="codigo">Codigo: <span id="cod"></span></div>
    <div class="descricao" id="desc"></div>
    <div class="imgwrap"><img id="img" alt="imagem do produto"></div>
    <div><span class="badge" id="badge"></span></div>
    <div class="acoes">
      <button class="btn-no" onclick="validar('reprovado')">✗ NAO BATE</button>
      <button class="btn-ok" onclick="validar('aprovado')">✓ BATE</button>
    </div>
    <div class="nav">
      <button id="prev" onclick="ir(-1)">← Anterior</button>
      <span class="pos" id="pos"></span>
      <button id="next" onclick="ir(1)">Proximo →</button>
    </div>
    <div class="dica">Atalhos: <b>1</b> ou <b>←seta esq</b> = nao bate · <b>2</b> ou <b>→seta dir</b> = bate · <b>Enter</b> = proximo</div>
  </div>
  <div class="vazio" id="vazio" style="display:none">Nenhum produto neste filtro. 🎉</div>
</main>

<script>
let lista = [];
let idx = 0;

async function carregarStats(){
  const s = await (await fetch('/api/stats')).json();
  document.getElementById('stats').innerHTML =
    `<span class="pill">Total c/ imagem: <b>${s.total}</b></span>`+
    `<span class="pill ok">Bate: <b>${s.aprovado}</b></span>`+
    `<span class="pill no">Nao bate: <b>${s.reprovado}</b></span>`+
    `<span class="pill pe">Pendentes: <b>${s.pendente}</b></span>`;
}

async function carregarLista(){
  const f = document.getElementById('filtro').value;
  const ordem = window.ORDEM || 'asc';
  lista = await (await fetch('/api/lista?status='+f+'&ordem='+ordem)).json();
  idx = 0;
  render();
  carregarStats();
}

function render(){
  const card = document.getElementById('card');
  const vazio = document.getElementById('vazio');
  if(lista.length === 0){ card.style.display='none'; vazio.style.display='block'; return; }
  vazio.style.display='none'; card.style.display='block';
  if(idx < 0) idx = 0;
  if(idx >= lista.length) idx = lista.length-1;
  const p = lista[idx];
  document.getElementById('cod').textContent = p.pro_codigo;
  document.getElementById('desc').textContent = p.pro_descricao || '(sem descricao)';
  document.getElementById('img').src = '/imagens/' + encodeURIComponent(p.imagem);
  const badge = document.getElementById('badge');
  const map = {aprovado:['ok','✓ BATE'], reprovado:['no','✗ NAO BATE'], pendente:['pe','• Pendente']};
  const [cls,txt] = map[p.status] || map.pendente;
  badge.className = 'badge '+cls; badge.textContent = txt;
  document.getElementById('pos').textContent = (idx+1)+' de '+lista.length;
  document.getElementById('prev').disabled = idx===0;
  document.getElementById('next').disabled = idx===lista.length-1;
}

function ir(d){ idx += d; render(); }

async function validar(status){
  const p = lista[idx];
  if(!p) return;
  let link = null;
  if(status === 'reprovado'){
    link = prompt('Cole o link da imagem correta (opcional):', p.link || '');
    if(link === null) return; // cancelou: nao valida
    link = link.trim() || null;
  }
  await fetch('/api/validar', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({pro_codigo:p.pro_codigo, status, link})});
  p.status = status;
  p.link = link;
  carregarStats();
  // se o item validado nao pertence mais ao filtro atual, ele some da tela na hora
  const f = document.getElementById('filtro').value;
  if(f !== 'todos' && p.status !== f){ lista.splice(idx,1); render(); }
  else { if(idx < lista.length-1){ idx++; } render(); }
}

document.getElementById('filtro').addEventListener('change', carregarLista);
document.getElementById('btn-sair').addEventListener('click', ()=>{
  fetch('/logout', {method:'POST'}).then(()=>{ location.href='/login'; });
});
document.addEventListener('keydown', e=>{
  if(e.key==='1' || e.key==='ArrowLeft' && e.shiftKey){ validar('reprovado'); }
  else if(e.key==='2' || e.key==='ArrowRight' && e.shiftKey){ validar('aprovado'); }
  else if(e.key==='ArrowLeft'){ ir(-1); }
  else if(e.key==='ArrowRight'){ ir(1); }
  else if(e.key==='Enter'){ ir(1); }
});

if(window.VALIDADOR){
  document.getElementById('quem').textContent = '— '+window.VALIDADOR;
  document.title = 'Validador — '+window.VALIDADOR;
}
if(window.USUARIO_LOGADO){
  document.getElementById('usuario-logado').textContent = window.USUARIO_LOGADO;
}
carregarLista();
</script>
</body>
</html>
"""


# ----------------------------------------------------------------------------
# Inicializacao
# ----------------------------------------------------------------------------
if __name__ == "__main__":
    if not DATABASE:
        raise SystemExit(
            "ERRO: variavel DATABASE vazia no .env.\n"
            "Preencha com a connection string do PostgreSQL, ex:\n"
            "DATABASE=postgresql://usuario:senha@host:5432/nome_do_banco"
        )
    print("Inicializando banco de dados...")
    init_db()
    print("Carregando produtos do Excel (isso pode levar alguns segundos)...")
    novos = carregar_excel()
    print(f"OK. {novos} novos produtos adicionados.")
    print("\nAbra no navegador:  http://127.0.0.1:5000\n")
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "5000"))
    app.run(debug=False, host=host, port=port)
